"""
ETMall 大平台貼單 CSV 轉 XLSX 腳本

功能：
- 讀取步驟1輸出的最新CSV檔案
- 轉換為XLSX格式
- 按月份分頁
- 標題翻譯回中文版
- 輸出到 data_processed 目錄

輸入：temp/etmall/01_etmall_platform_orders_merged_*.csv
輸出：data_processed/etmall/etmall_platform_orders_monthly.xlsx
"""

import sys
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def setup_logging() -> None:
    """設定日誌"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )


def find_latest_csv_file(temp_dir: Path) -> Optional[Path]:
    """尋找最新的CSV檔案"""
    pattern = "01_etmall_platform_orders_merged_*.csv"
    files = list(temp_dir.glob(pattern))
    
    if not files:
        logging.warning(f'在 {temp_dir} 中找不到 {pattern} 檔案')
        return None
    
    # 按修改時間排序，最新的在前
    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    latest_file = files[0]
    
    logging.info(f'找到最新CSV檔案：{latest_file.name}')
    logging.info(f'修改時間：{datetime.fromtimestamp(latest_file.stat().st_mtime)}')
    
    return latest_file


def get_column_mapping() -> Dict[str, str]:
    """取得欄位名稱對應表（英文到中文）"""
    return {
        'platform': '平台',
        'order_date': '訂單日期',
        'order_sn': '訂單編號',
        'item_no': '項目編號',
        'order_line_uid': '訂單行唯一識別碼',
        'shipping_carrier': '物流公司',
        'shipping_expected_date': '預期送達日',
        'shipping_expected_time': '預期送達時段',
        'package_count': '包裹數量',
        'extra_shipping_fee': '額外運費',
        'customer_name': '客戶姓名',
        'shipping_zipcode': '郵遞區號',
        'shipping_address_detail': '詳細地址',
        'customer_phone': '客戶手機',
        'customer_tel_day': '客戶日間電話',
        'customer_tel_night': '客戶夜間電話',
        'note_1': '備註1',
        'note_2': '備註2',
        'product_sale_id': '商品銷售編號',
        'color': '顏色',
        'product_name_platform': '平台商品名稱',
        'quantity': '數量',
        'unit_price': '單價',
        'order_amount': '訂單金額',
        'platform_reconciliation_cost': '平台對帳成本',
        'supplier_cost': '供應商成本',
        'platform_commission_rate': '平台佣金率',
        'profit': '利潤',
        'cod_amount': '代收金額',
        'invoice_no': '發票編號',
        'reconciliation_item_no': '對帳項目編號',
        'is_gift': '是否贈品',
        'bank_account_last5': '銀行帳戶末五碼'
    }


def format_worksheet(ws: Worksheet, title: str) -> None:
    """格式化工作表"""
    # 設定欄寬
    column_widths = {
        'A': 8,   # 平台
        'B': 12,  # 訂單日期
        'C': 15,  # 訂單編號
        'D': 10,  # 項目編號
        'E': 20,  # 訂單行唯一識別碼
        'F': 12,  # 物流公司
        'G': 12,  # 預期送達日
        'H': 15,  # 預期送達時段
        'I': 10,  # 包裹數量
        'J': 12,  # 額外運費
        'K': 15,  # 客戶姓名
        'L': 10,  # 郵遞區號
        'M': 30,  # 詳細地址
        'N': 15,  # 客戶手機
        'O': 15,  # 客戶日間電話
        'P': 15,  # 客戶夜間電話
        'Q': 20,  # 備註1
        'R': 20,  # 備註2
        'S': 15,  # 商品銷售編號
        'T': 10,  # 顏色
        'U': 10,  # 款式
        'V': 30,  # 平台商品名稱
        'W': 8,   # 數量
        'X': 12,  # 單價
        'Y': 12,  # 訂單金額
        'Z': 12,  # 對帳金額
        'AA': 12, # 平台成本
        'AB': 15, # 平台佣金率
        'AC': 10, # 利潤
        'AD': 12, # 代收金額
        'AE': 15, # 發票編號
        'AF': 15, # 對帳項目編號
        'AG': 10, # 是否贈品
        'AH': 10, # 郵遞區號
        'AI': 30, # 詳細地址
        'AJ': 12  # 銀行帳戶末五碼
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 設定標題行樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 設定邊框樣式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 套用標題行樣式
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    # 設定資料行邊框
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border


def create_monthly_worksheets(wb: openpyxl.Workbook, df: pd.DataFrame, column_mapping: Dict[str, str]) -> None:
    """建立按月份分頁的工作表"""
    # 確保 order_date 是日期格式
    df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce')
    
    # 移除無效日期的資料
    df_valid = df.dropna(subset=['order_date'])
    
    if df_valid.empty:
        logging.warning('沒有有效的日期資料，無法建立月份分頁')
        return
    
    # 按月份分組
    df_valid['year_month'] = df_valid['order_date'].dt.to_period('M')
    monthly_groups = df_valid.groupby('year_month')
    
    logging.info(f'開始建立月份分頁，共 {len(monthly_groups)} 個月')
    
    # 建立每個月份的工作表
    for year_month, month_data in monthly_groups:
        # 工作表名稱格式：YYYY年MM月
        sheet_name = f"{year_month.year}年{year_month.month:02d}月"
        
        # 如果工作表名稱太長，使用縮短版本
        if len(sheet_name) > 31:
            sheet_name = f"{year_month.year}-{year_month.month:02d}"
        
        # 建立工作表
        ws = wb.create_sheet(title=sheet_name)
        
        # 重新排序資料（按日期和訂單編號）
        month_data_sorted = month_data.sort_values(['order_date', 'order_sn'], ascending=[True, True])
        
        # 準備標題行
        headers = []
        for col in month_data_sorted.columns:
            if col in column_mapping:
                headers.append(column_mapping[col])
            else:
                headers.append(col)
        
        # 寫入標題行
        ws.append(headers)
        
        # 寫入資料
        for _, row in month_data_sorted.iterrows():
            # 清理資料，將特殊類型轉換為字串
            cleaned_row = []
            for value in row.tolist():
                if pd.isna(value):
                    cleaned_row.append('')
                elif isinstance(value, pd.Period):
                    cleaned_row.append(str(value))
                elif isinstance(value, pd.Timestamp):
                    cleaned_row.append(value.strftime('%Y-%m-%d'))
                else:
                    cleaned_row.append(str(value))
            ws.append(cleaned_row)
        
        # 格式化工作表
        format_worksheet(ws, sheet_name)
        
        logging.info(f'建立工作表：{sheet_name}，資料行數：{len(month_data_sorted)}')


def main() -> None:
    """主函數"""
    setup_logging()
    
    # 取得專案根目錄
    project_root = Path(__file__).resolve().parents[2]
    temp_dir = project_root / 'temp' / 'etmall'
    output_dir = project_root / 'data_processed' / 'etmall'
    
    logging.info(f'專案根目錄：{project_root}')
    logging.info(f'來源目錄：{temp_dir}')
    logging.info(f'輸出目錄：{output_dir}')
    
    # 尋找最新的CSV檔案
    latest_csv = find_latest_csv_file(temp_dir)
    if not latest_csv:
        logging.error('找不到CSV檔案，程式結束')
        sys.exit(1)
    
    # 讀取CSV檔案
    try:
        logging.info(f'開始讀取CSV檔案：{latest_csv.name}')
        
        # 指定特定欄位為字串類型，確保格式不被改變
        dtype_dict = {
            'package_count': str,      # 包裹數量
            'shipping_zipcode': str,   # 郵遞區號
            'customer_phone': str,     # 客戶手機
            'customer_tel_day': str,   # 客戶日間電話
            'customer_tel_night': str, # 客戶夜間電話
            'product_sale_id': str,    # 商品銷售編號
            'color': str,              # 顏色
            'invoice_no': str,         # 發票編號
            'reconciliation_item_no': str, # 對帳項目編號
            'order_sn': str,           # 訂單編號
            'item_no': str,            # 項目編號
            'order_line_uid': str      # 訂單行唯一識別碼
        }
        
        df = pd.read_csv(latest_csv, encoding='utf-8-sig', dtype=dtype_dict)
        
        # 處理特定欄位的格式問題
        # 1. 包裹數量：去除小數點
        if 'package_count' in df.columns:
            df['package_count'] = df['package_count'].astype(str).str.replace(r'\.0+$', '', regex=True)
            df['package_count'] = df['package_count'].replace(['nan', 'None', 'NULL'], '')
        
        # 2. 郵遞區號：去除小數點
        if 'shipping_zipcode' in df.columns:
            df['shipping_zipcode'] = df['shipping_zipcode'].astype(str).str.replace(r'\.0+$', '', regex=True)
            df['shipping_zipcode'] = df['shipping_zipcode'].replace(['nan', 'None', 'NULL'], '')
        
        # 3. 電話號碼：確保9開頭9碼的號碼有前導0
        phone_columns = ['customer_phone', 'customer_tel_day', 'customer_tel_night']
        for col in phone_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(r'\.0+$', '', regex=True)
                # 處理9開頭9碼的號碼
                df[col] = df[col].apply(lambda x: '0' + x if len(x) == 9 and x.startswith('9') and x.isdigit() else x)
                df[col] = df[col].replace(['nan', 'None', 'NULL'], '')
        
        logging.info(f'CSV檔案讀取完成，資料行數：{len(df)}，欄位數：{len(df.columns)}')
    except Exception as e:
        logging.error(f'讀取CSV檔案時發生錯誤：{e}')
        sys.exit(1)
    
    # 取得欄位對應表
    column_mapping = get_column_mapping()
    
    # 建立新的工作簿
    wb = openpyxl.Workbook()
    
    # 移除預設工作表
    wb.remove(wb.active)
    
    # 建立月份分頁
    create_monthly_worksheets(wb, df, column_mapping)
    
    # 確保輸出目錄存在
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 生成輸出檔案名稱
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'etmall_platform_orders_monthly_{timestamp}.xlsx'
    output_path = output_dir / output_filename
    
    # 儲存XLSX檔案
    try:
        wb.save(output_path)
        logging.info(f'XLSX檔案儲存完成：{output_path}')
    except Exception as e:
        logging.error(f'儲存XLSX檔案時發生錯誤：{e}')
        sys.exit(1)
    
    # 輸出處理結果
    logging.info(f'\n=== 轉換完成 ===')
    logging.info(f'輸入檔案：{latest_csv.name}')
    logging.info(f'輸出檔案：{output_filename}')
    logging.info(f'輸出位置：{output_path}')
    
    # 顯示工作表資訊
    sheet_names = wb.sheetnames
    logging.info(f'建立的工作表：{", ".join(sheet_names)}')


if __name__ == '__main__':
    main()
