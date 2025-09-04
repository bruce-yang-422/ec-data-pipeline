#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
資料日期檢查腳本

主要功能：
- 檢查 data_processed/merged 目錄下最新的 etmall_orders_product_enriched_*.csv 檔案
- 分析 order_date 的資料完整性，找出遺漏的日期
- 檢查有 order_date 和 order_sn 但沒有 product_name_platform 的資料
- 生成詳細的日期分析報告

Authors: 楊翔志 & AI Collective
Studio: tranquility-base
"""

import os
import sys
import pandas as pd
import glob
import logging
from datetime import datetime, timedelta
from typing import List, Set, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ✅ 將專案根目錄加入 sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

def setup_logging():
    """設定日誌系統"""
    # 取得腳本所在目錄
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 確保腳本目錄存在
    os.makedirs(script_dir, exist_ok=True)
    
    # 刪除舊的報告檔案（.txt 和 .xlsx）
    old_txt_files = glob.glob(os.path.join(script_dir, "check_data_dates_*.txt"))
    old_xlsx_files = glob.glob(os.path.join(script_dir, "check_data_dates_*.xlsx"))
    
    for old_file in old_txt_files + old_xlsx_files:
        try:
            os.remove(old_file)
            print(f"[INFO] 已刪除舊報告檔案：{os.path.basename(old_file)}")
        except Exception as e:
            print(f"[WARN] 刪除舊報告檔案失敗 {os.path.basename(old_file)}: {e}")
    
    # 設定報告檔案名稱
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = os.path.join(script_dir, f"check_data_dates_{timestamp}.txt")
    excel_file = os.path.join(script_dir, f"check_data_dates_{timestamp}.xlsx")
    
    # 設定日誌格式
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(report_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.excel_file = excel_file  # 將 Excel 檔案路徑附加到 logger 物件
    return logger

def get_latest_etmall_file():
    """取得最新的 ETMall 產品資料豐富化 CSV 檔案"""
    current_dir = os.getcwd()
    
    # 如果當前目錄是專案根目錄
    if os.path.basename(current_dir) == "ec-data-pipeline":
        pattern = "data_processed/merged/etmall_orders_product_enriched_*.csv"
    else:
        # 如果從其他目錄執行，使用絕對路徑
        pattern = os.path.join(current_dir, "data_processed/merged/etmall_orders_product_enriched_*.csv")
    
    files = glob.glob(pattern)
    
    if not files:
        raise FileNotFoundError(f"找不到符合模式的檔案: {pattern}")
    
    # 按修改時間排序，取得最新的檔案
    latest_file = max(files, key=os.path.getmtime)
    return latest_file

def analyze_date_gaps(df: pd.DataFrame, logger: logging.Logger) -> Dict[str, Any]:
    """分析日期間隔，找出遺漏的日期"""
    logger.info("=== 分析日期間隔 ===")
    
    # 轉換 order_date 為 datetime
    df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce')
    
    # 移除無效的日期
    valid_dates = df['order_date'].dropna()
    
    if len(valid_dates) == 0:
        logger.warning("沒有找到有效的 order_date 資料")
        return {"missing_dates": [], "date_range": None, "total_days": 0, "missing_count": 0}
    
    # 取得日期範圍
    min_date = valid_dates.min()
    max_date = valid_dates.max()
    
    logger.info(f"資料日期範圍: {min_date.strftime('%Y-%m-%d')} ~ {max_date.strftime('%Y-%m-%d')}")
    
    # 生成完整的日期範圍
    date_range = pd.date_range(start=min_date, end=max_date, freq='D')
    
    # 找出實際存在的日期
    existing_dates = set(valid_dates.dt.date)
    expected_dates = set(date_range.date)
    
    # 找出遺漏的日期
    missing_dates = sorted(expected_dates - existing_dates)
    
    logger.info(f"總天數: {len(date_range)}")
    logger.info(f"有資料的天數: {len(existing_dates)}")
    logger.info(f"遺漏的天數: {len(missing_dates)}")
    
    if missing_dates:
        logger.info("遺漏的日期:")
        for date in missing_dates:
            logger.info(f"  - {date}")
    else:
        logger.info("✅ 沒有遺漏的日期")
    
    return {
        "missing_dates": missing_dates,
        "date_range": (min_date, max_date),
        "total_days": len(date_range),
        "missing_count": len(missing_dates),
        "existing_dates": existing_dates
    }

def analyze_missing_shipping_data(df: pd.DataFrame, logger: logging.Logger) -> Dict[str, Any]:
    """分析有 order_date 和 order_sn 但沒有 customer_name 的資料（缺少出貨資料）"""
    logger.info("\n=== 分析缺少出貨資料 ===")
    
    # 檢查欄位是否存在
    required_columns = ['order_date', 'order_sn', 'customer_name', 'shipping_status']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        logger.error(f"缺少必要欄位: {missing_columns}")
        return {"missing_data": pd.DataFrame(), "date_summary": {}}
    
    logger.info("檢查缺少出貨資料（customer_name 為空，但排除已取消和銷退的訂單）")
    
    # 轉換 order_date 為 datetime
    df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce')
    
    # 找出有 order_date 和 order_sn 但沒有 customer_name 的資料（缺少出貨資料）
    # 條件：order_date 不為空 AND order_sn 不為空 AND (customer_name 為空 OR 為 NaN) AND shipping_status 不包含"取消"或"銷退"
    missing_shipping_mask = (
        df['order_date'].notna() & 
        df['order_sn'].notna() & 
        (df['customer_name'].isna() | (df['customer_name'] == '') | (df['customer_name'] == 'nan')) &
        (~df['shipping_status'].str.contains('取消|銷退', na=False))  # 排除包含"取消"或"銷退"的訂單
    )
    
    missing_data = df[missing_shipping_mask].copy()
    
    logger.info(f"缺少出貨資料的筆數: {len(missing_data)}")
    
    if len(missing_data) > 0:
        # 按日期分組統計（以去除重複的 order_sn 為單位）
        missing_data['date_str'] = missing_data['order_date'].dt.strftime('%Y-%m-%d')
        date_summary = missing_data.groupby('date_str').agg({
            'order_sn': 'nunique'  # 使用 nunique 來計算不重複的 order_sn 數量
        }).rename(columns={'order_sn': 'unique_order_count'})
        
        logger.info("按日期統計缺少出貨資料:")
        for date, row in date_summary.iterrows():
            logger.info(f"  {date}: {row['unique_order_count']} 筆訂單")
        
        # 顯示日期範圍
        if len(missing_data) > 0:
            min_date = missing_data['order_date'].min()
            max_date = missing_data['order_date'].max()
            logger.info(f"缺少出貨資料的日期範圍: {min_date.strftime('%Y-%m-%d')} ~ {max_date.strftime('%Y-%m-%d')}")
        
        # 顯示一些範例資料
        logger.info("範例缺少出貨資料:")
        sample_cols = ['order_date', 'order_sn', 'customer_name', 'shipping_status']
        available_cols = [col for col in sample_cols if col in missing_data.columns]
        
        for idx, row in missing_data[available_cols].head(5).iterrows():
            # 轉換 pandas 類型為 Python 原生類型
            row_dict = {}
            for col in available_cols:
                value = row[col]
                if pd.isna(value):
                    row_dict[col] = None
                elif hasattr(value, 'strftime'):  # datetime 類型
                    row_dict[col] = value.strftime('%Y-%m-%d')
                else:
                    row_dict[col] = str(value)
            logger.info(f"  {row_dict}")
    else:
        logger.info("✅ 沒有發現缺少出貨資料")
        date_summary = {}
    
    return {
        "missing_data": missing_data,
        "date_summary": date_summary
    }

def generate_summary_report(date_analysis: Dict[str, Any], shipping_analysis: Dict[str, Any], logger: logging.Logger):
    """生成總結報告"""
    logger.info("\n" + "="*60)
    logger.info("📊 資料完整性檢查總結報告")
    logger.info("="*60)
    
    # 日期完整性報告
    logger.info("\n📅 日期完整性分析:")
    if date_analysis["date_range"]:
        min_date, max_date = date_analysis["date_range"]
        logger.info(f"  資料日期範圍: {min_date.strftime('%Y-%m-%d')} ~ {max_date.strftime('%Y-%m-%d')}")
        logger.info(f"  總天數: {date_analysis['total_days']}")
        logger.info(f"  有資料的天數: {date_analysis['total_days'] - date_analysis['missing_count']}")
        logger.info(f"  遺漏的天數: {date_analysis['missing_count']}")
        
        if date_analysis['missing_count'] > 0:
            logger.info(f"  完整性: {((date_analysis['total_days'] - date_analysis['missing_count']) / date_analysis['total_days'] * 100):.1f}%")
        else:
            logger.info("  完整性: 100%")
    
    # 出貨資料完整性報告
    logger.info("\n📦 出貨資料完整性分析:")
    missing_data = shipping_analysis["missing_data"]
    if len(missing_data) > 0:
        logger.info(f"  缺少出貨資料的記錄數: {len(missing_data)}")
        unique_orders = missing_data['order_sn'].nunique()
        logger.info(f"  缺少出貨資料的訂單數: {unique_orders}")
        if not shipping_analysis["date_summary"].empty:
            date_count = len(shipping_analysis["date_summary"])
            logger.info(f"  影響的日期數: {date_count}")
            
            # 顯示具體的日期列表
            logger.info("  缺少出貨資料的日期:")
            for date, row in shipping_analysis["date_summary"].iterrows():
                logger.info(f"    - {date}: {row['unique_order_count']} 筆訂單")
    else:
        logger.info("  出貨資料完整性: 100%")
    
    # 建議
    logger.info("\n💡 建議:")
    if date_analysis['missing_count'] > 0:
        logger.info("  - 檢查遺漏日期的原始資料檔案")
        logger.info("  - 確認資料處理流程是否完整")
    
    if len(missing_data) > 0:
        logger.info("  - 檢查上述日期的出貨資料收集是否完整")
        logger.info("  - 確認客戶名稱欄位的資料來源")
        logger.info("  - 補齊缺少的出貨資料")
    
    if date_analysis['missing_count'] == 0 and len(missing_data) == 0:
        logger.info("  - ✅ 資料完整性良好，無需特別處理")

def generate_excel_report(date_analysis: Dict[str, Any], shipping_analysis: Dict[str, Any], logger: logging.Logger):
    """生成 Excel 格式的純報告"""
    try:
        # 創建新的工作簿
        wb = openpyxl.Workbook()
        
        # 移除預設的工作表
        wb.remove(wb.active)
        
        # 1. 建立摘要工作表
        summary_ws = wb.create_sheet("資料完整性摘要")
        
        # 設定標題樣式
        title_font = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='微軟正黑體', size=12, bold=True)
        data_font = Font(name='微軟正黑體', size=11)
        
        # 摘要標題
        summary_ws['A1'] = 'ETMall 資料完整性檢查報告'
        summary_ws['A1'].font = title_font
        summary_ws['A1'].fill = title_fill
        summary_ws.merge_cells('A1:D1')
        
        # 日期完整性摘要
        row = 3
        summary_ws[f'A{row}'] = '📅 日期完整性分析'
        summary_ws[f'A{row}'].font = header_font
        
        if date_analysis["date_range"]:
            min_date, max_date = date_analysis["date_range"]
            summary_ws[f'A{row+1}'] = '資料日期範圍'
            summary_ws[f'B{row+1}'] = f"{min_date.strftime('%Y-%m-%d')} ~ {max_date.strftime('%Y-%m-%d')}"
            summary_ws[f'A{row+2}'] = '總天數'
            summary_ws[f'B{row+2}'] = date_analysis['total_days']
            summary_ws[f'A{row+3}'] = '有資料的天數'
            summary_ws[f'B{row+3}'] = date_analysis['total_days'] - date_analysis['missing_count']
            summary_ws[f'A{row+4}'] = '遺漏的天數'
            summary_ws[f'B{row+4}'] = date_analysis['missing_count']
            summary_ws[f'A{row+5}'] = '完整性'
            if date_analysis['missing_count'] > 0:
                completeness = ((date_analysis['total_days'] - date_analysis['missing_count']) / date_analysis['total_days'] * 100)
                summary_ws[f'B{row+5}'] = f"{completeness:.1f}%"
            else:
                summary_ws[f'B{row+5}'] = "100%"
        
        # 出貨資料完整性摘要
        row += 7
        summary_ws[f'A{row}'] = '📦 出貨資料完整性分析'
        summary_ws[f'A{row}'].font = header_font
        
        missing_data = shipping_analysis["missing_data"]
        summary_ws[f'A{row+1}'] = '缺少出貨資料的記錄數'
        summary_ws[f'B{row+1}'] = len(missing_data)
        
        if len(missing_data) > 0:
            # 計算不重複的訂單數
            unique_orders = missing_data['order_sn'].nunique()
            summary_ws[f'A{row+2}'] = '缺少出貨資料的訂單數'
            summary_ws[f'B{row+2}'] = unique_orders
            summary_ws[f'A{row+3}'] = '影響的日期數'
            summary_ws[f'B{row+3}'] = len(shipping_analysis["date_summary"])
        else:
            summary_ws[f'A{row+2}'] = '出貨資料完整性'
            summary_ws[f'B{row+2}'] = '100%'
        
        # 設定字體
        for row in summary_ws.iter_rows():
            for cell in row:
                if cell.value and not cell.font.bold:
                    cell.font = data_font
        
        # 2. 建立缺少出貨資料的詳細工作表
        if len(missing_data) > 0:
            detail_ws = wb.create_sheet("缺少出貨資料明細")
            
            # 標題
            detail_ws['A1'] = '缺少出貨資料的日期明細'
            detail_ws['A1'].font = title_font
            detail_ws['A1'].fill = title_fill
            detail_ws.merge_cells('A1:D1')
            
            # 表頭
            detail_ws['A3'] = '日期'
            detail_ws['B3'] = '星期'
            detail_ws['C3'] = '缺少訂單數'
            detail_ws['D3'] = '備註'
            detail_ws['A3'].font = header_font
            detail_ws['B3'].font = header_font
            detail_ws['C3'].font = header_font
            detail_ws['D3'].font = header_font
            
            # 填入資料
            row = 4
            for date, row_data in shipping_analysis["date_summary"].iterrows():
                # 轉換日期字串為 datetime 物件以取得星期
                date_obj = pd.to_datetime(date)
                weekday = date_obj.strftime('%A')  # 英文星期
                weekday_cn = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日'][date_obj.weekday()]
                
                detail_ws[f'A{row}'] = date
                detail_ws[f'B{row}'] = weekday_cn
                detail_ws[f'C{row}'] = row_data['unique_order_count']
                detail_ws[f'D{row}'] = '需要補齊出貨資料'
                detail_ws[f'A{row}'].font = data_font
                detail_ws[f'B{row}'].font = data_font
                detail_ws[f'C{row}'].font = data_font
                detail_ws[f'D{row}'].font = data_font
                row += 1
            
            # 設定欄寬
            detail_ws.column_dimensions['A'].width = 15
            detail_ws.column_dimensions['B'].width = 10
            detail_ws.column_dimensions['C'].width = 12
            detail_ws.column_dimensions['D'].width = 25
        

        
        # 設定摘要工作表的欄寬
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 25
        
        # 儲存檔案
        wb.save(logger.excel_file)
        logger.info(f"✅ Excel 報告已生成：{os.path.basename(logger.excel_file)}")
        
    except Exception as e:
        logger.error(f"❌ 生成 Excel 報告失敗：{str(e)}")

def main():
    """主程式"""
    logger = setup_logging()
    
    try:
        # 取得最新的 ETMall 檔案
        csv_file = get_latest_etmall_file()
        logger.info(f"使用檔案: {csv_file}")
        
        # 讀取 CSV 檔案
        logger.info("讀取 CSV 檔案...")
        df = pd.read_csv(csv_file, dtype=str, keep_default_na=False)
        logger.info(f"CSV 檔案筆數: {len(df)}")
        logger.info(f"CSV 檔案欄位數: {len(df.columns)}")
        
        # 顯示基本資訊
        logger.info("\n=== 檔案基本資訊 ===")
        logger.info(f"檔案路徑: {csv_file}")
        logger.info(f"檔案大小: {os.path.getsize(csv_file) / 1024 / 1024:.2f} MB")
        logger.info(f"最後修改時間: {datetime.fromtimestamp(os.path.getmtime(csv_file)).strftime('%Y-%m-%d %H:%M:%S')}")
        
        # 分析日期間隔
        date_analysis = analyze_date_gaps(df, logger)
        
        # 分析缺少出貨資料
        shipping_analysis = analyze_missing_shipping_data(df, logger)
        
        # 生成總結報告
        generate_summary_report(date_analysis, shipping_analysis, logger)
        
        # 生成 Excel 報告
        generate_excel_report(date_analysis, shipping_analysis, logger)
        
        logger.info("\n✅ 資料日期檢查完成")
        
    except Exception as e:
        logger.error(f"❌ 執行失敗: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
